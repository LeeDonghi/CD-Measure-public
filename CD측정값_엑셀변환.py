# -*- coding: utf-8 -*-
# ============================================================
#  CD 측정 사진 → 엑셀 자동 변환 프로그램
# ------------------------------------------------------------
#  실행하면 GUI 창이 뜸:
#   1) 공정(RDL/PI/Bump/Cu Post) 선택
#   2) CD 종류(Pad CD/Via CD/UBM CD/Cu Post CD) + (RDL이면 Line&Space도) Item별
#      측정 포인트 수·Target 값 입력, Overlay 포인트 수 입력
#      (Tolerance는 공정별 기준표에서 Target 값으로 자동 계산됨)
#   3) 사진 폴더 선택 → 자동으로 읽어서 같은 폴더에 "측정결과.xlsx" 생성
#
#  동작 방식:
#   - 사진마다 흰색 측정값 글자가 있으면 사용, 없으면 "측정 전" 사진이라 버림
#   - 남은 값 사진은 값 개수로 종류를 판별: 1개=CD류, 2개=L/S류, 4개=overlay
#   - 시간 순서로 3장씩 묶어 Point 하나로 처리 (overlay/CD/L/S 한 장씩)
#   - 측정값마다 입력받은 Item들 중 제일 가까운 Target을 찾아 편차를 계산
#     (사진 순서로 Item을 딱 맞추면 순서가 뒤섞일 때 위험하므로, 값이
#      제일 가까운 Item에 매칭하는 방식을 씀)
#   - USL/LSL(=Target±Tolerance) 벗어난 값은 "스펙이탈"에 표시하고 셀 색칠
#   - Item별 평균/표준편차/Cpk를 "Cpk요약" 시트에 정리 + Cpk 막대그래프
#   - OCR 신뢰도가 낮은 값은 "확인필요" 열에 표시
# ============================================================

import pandas as pd
import glob
import json
import os
import unicodedata
import sys
import re
import csv
import datetime
import subprocess
import time
import webbrowser
from collections import Counter
from pathlib import Path
from tkinter import messagebox
from openpyxl.styles import PatternFill, Font
from openpyxl.chart import BarChart, LineChart, Reference

from ocr_core import load_templates, read_image_ex, read_image_attempts, MIN_LINE_WIDTH
from plan_gui import collect_plan_and_folder
from measurement_plan import (match_single_target, match_ls_targets, summarize_matching,
                              compute_stats, display_item_name, find_ocr_outliers,
                              compute_item_breakdown)
from overlay_analysis import assign_directions, compute_overlay_metrics
from ls_brightness import verify_ls_brightness
from html_report import build_report

# 읽을 수 있는 사진 형식. 계측기 설정/버전에 따라 저장 형식이 달라짐.
# (.jpg는 손실 압축이라 글자 가장자리가 살짝 뭉개지는데, 실제 사진으로
#  검증해 보니 판독 결과는 .bmp와 동일했음)
IMAGE_EXTENSIONS = ('.bmp', '.png', '.jpg', '.jpeg')

KIND_LABEL_BY_COUNT = {1: 'CD', 2: 'L/S', 4: 'overlay'}

# 계측기가 두 측정줄을 겹쳐 그려서 값을 못 읽은 사진에 붙이는 표시.
# 로그에서 이 표시를 세어 완료 안내창에 알려줌.
UNREAD_MARK = '[겹침]'


def unread_by_file(log):
    """log에서 "[겹침] 파일명: ..." 줄을 찾아 {파일명: True}로 뽑아준다.
       read_measured가 남기는 형식(줄 앞머리 "  [겹침] 이름.png: ...")에
       맞춘 것 - measurement_plan.compute_item_breakdown이 log 문자열
       형식을 몰라도 되도록 여기서 구조화해서 넘기려고 만듦(2026-08-12)."""
    result = {}
    for line in log:
        if UNREAD_MARK not in line:
            continue
        after = line.split(UNREAD_MARK, 1)[1].strip()
        name = after.split(':', 1)[0].strip()
        result[name] = True
    return result

# 실행 이력(요약 통계) CSV — 프로그램 폴더에 고정. 측정 사진 폴더(Lot마다 다름)에
# 두면 폴더별로 흩어져서 "여러 번 실행을 모아 추세를 보는" 목적에 안 맞으므로,
# 여러 Lot을 넘나들며 한곳에서 누적되도록 스크립트 자신의 위치를 기준으로 잡음.
#
# CD_MEASURE_DATA_DIR 환경변수가 있으면 그 폴더에 둠. 웹 버전을 Docker로 배포할
# 때 쓰는 값 — 컨테이너 "안"에 쌓으면 이미지를 새로 만들 때마다 누적 이력이
# 통째로 사라지므로, 바깥(서버) 폴더를 연결해서 그쪽에 쌓게 하려는 것.
# 환경변수가 없으면 예전과 완전히 동일하게 동작함(데스크탑 실행에는 영향 없음).
DATA_DIR = os.environ.get('CD_MEASURE_DATA_DIR') or os.path.dirname(os.path.abspath(__file__))
RUN_HISTORY_CSV = os.path.join(DATA_DIR, '실행이력.csv')
# 컬럼을 새로 넣을 때는 반드시 **맨 뒤에** 붙일 것.
# 중간에 끼워 넣으면 이미 쌓인 줄들의 값이 한 칸씩 밀려 과거 기록이 통째로
# 어긋난다(실행이력.csv는 여러 Lot에 걸친 누적 통계라 다시 만들 수 없음).
# 뒤에 붙이면 옛 줄은 그 칸이 빈 값으로 읽혀 그대로 쓸 수 있다.
#   실행자 : 웹에서 로그인한 아이디. 데스크탑 실행은 로그인 개념이 없어 빈 값 (2026-08-04 추가)
#   소요초 : 사진 읽기~저장까지 걸린 시간(초). 서버가 느려지는 추세를 보려는 것
#   구분   : 데모 실행이면 '데모', 실제 실행이면 빈 값 (2026-08-18 추가 - 공개
#            데모 실행이 실제 공정 통계에 섞이지 않도록 구분)
RUN_HISTORY_HEADER = [
    '실행시각', '폴더', '공정', '고객', '자재', 'Layer',
    '입력사진수', '측정사진수', '총측정값', 'Point수',
    '확인필요', '스펙이탈', '겹침사진수', '세트불일치Point수', '매칭불일치Item수',
    '실행자', '소요초', '구분',
]


def _migrate_run_history_header():
    """실행이력.csv에 컬럼을 새로 추가해도, 이미 있던 파일의 1번째 줄(헤더)은
       자동으로 안 바뀐다(save_run_log는 파일이 아예 없을 때만 헤더를 씀).
       2026-08-04에 실행자·소요초를 추가했을 때도 이 마이그레이션이 없어서
       헤더가 지금까지 옛날 상태로 남아있었음(2026-08-18에 발견).
       데이터 줄은 절대 건드리지 않고 헤더 줄만 안전하게 다시 쓴다 - 뒤에
       붙은 컬럼은 옛 줄에서 그냥 빈칸으로 읽히므로 문제없다.
       쓰기는 temp-file + os.replace()로 원자성 보장(Excel이 파일을 열고 있어도
       안전하게 실패하거나, 성공할 때만 기존 파일과 교체됨 - 부분 쓰기로 인한
       데이터 손상 방지)."""
    if not os.path.exists(RUN_HISTORY_CSV):
        return
    try:
        with open(RUN_HISTORY_CSV, 'r', newline='', encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
    except OSError:
        return
    if not rows or rows[0] == RUN_HISTORY_HEADER:
        return
    rows[0] = RUN_HISTORY_HEADER
    # 원자적 쓰기: 임시파일에 먼저 쓴 뒤 기존 파일과 교체(PermissionError 등이 발생해도
    # 기존 데이터는 안 건드려짐)
    try:
        tmp = RUN_HISTORY_CSV + '.tmp'
        with open(tmp, 'w', newline='', encoding='utf-8-sig') as f:
            csv.writer(f).writerows(rows)
        os.replace(tmp, RUN_HISTORY_CSV)
    except OSError:
        # Excel이 열고 있거나 권한 문제가 있으면 이번엔 건너뛰고 다음 실행에 재시도
        return


def display_kind(category, plan):
    if category == 'CD':
        return plan.cd_label
    if category == 'L/S':
        return 'Line&Space'
    if category == 'overlay':
        return 'Overlay'
    return category


# ------------------------------------------------------------
#  [핵심 기능] 사진들을 순서대로 읽어 Point 단위로 묶고, Target과 비교
# ------------------------------------------------------------
# 판독 근거를 남기는 곳. 글(JSON)과 줄 그림(crop)을 한 폴더에 같이 둔다.
#
# ⚠️ **사진과 달리 계속 남긴다.** 몇 달치가 쌓여야 "어느 밝기 기준이 안전한가",
# "2차 엔진이 필요한가" 같은 판단을 실측으로 할 수 있기 때문. 지금은 사진 몇
# 장의 관찰뿐이라 상수로 박을 근거가 부족함.
#
# 용량은 재봤을 때 문제가 안 됨 — 문제 있는 사진 10장짜리 실행에서 글 17KB +
# 그림 28장 364KB로, 사진 37.6MB의 1%였음(2026-08-05 실측). 그래서 그림도
# 같이 남긴다. 몇 달 뒤에 "그때 그 줄이 어떻게 생겼길래"를 눈으로 볼 수 있는
# 게 원인 분석에 결정적임.
#
# 서버 정리 스크립트(scripts/setup_server_backup.py)가 이 폴더를 지우지 않도록
# 예외로 빼두어야 함 — 안 그러면 7일 뒤 사진과 함께 사라짐.
OCR_EVIDENCE_DIR = 'ocr_evidence'
OCR_EVIDENCE_JSON = 'OCR근거.json'


def collect_ocr_evidence(run_dir, path, rows, line_count, templates):
    """판독에 문제가 있는 사진의 "왜 그렇게 읽혔는지"를 모아서 반환.

    2026-08-05에 9-1_22.png 한 장의 원인을 찾느라 서버에서 사진을 내려받고
    변환해서 눈으로 보는 데 시간을 다 썼음. 그때 필요했던 것이 정확히 이것 —
    밝기 기준별로 무슨 글자로 읽혔는지, 그리고 그 줄의 원본 그림.

    그림만 여기서 바로 저장하고, 글은 반환해서 실행 단위로 한 파일에 모은다.
    """
    from PIL import Image     # 근거를 남길 때만 필요해서 여기서 불러옴

    attempts = read_image_attempts(path, templates)
    name = os.path.splitext(os.path.basename(path))[0]

    # 줄 그림은 사람이 눈으로 봐야 하는 것이라, 기본 기준으로 찾은 줄만 저장.
    # 여백을 조금 둬서 위아래에 뭐가 붙었는지도 같이 보이게 함
    # (9-1_22는 글자 아래 밝은 자국이 붙은 게 원인이었는데, 딱 맞춰 자르면 안 보임).
    crop_dir = os.path.join(run_dir, OCR_EVIDENCE_DIR, name)
    os.makedirs(crop_dir, exist_ok=True)
    img = Image.open(path)
    base = next((a for a in attempts if a['method'] == 'template@170'), None)
    for i, (y0, y1, x0, x1) in enumerate(base['boxes'] if base else []):
        if (x1 - x0) < MIN_LINE_WIDTH:
            continue                      # 화면 상단 안내문구 등은 뺌
        img.crop((max(0, x0 - 8), max(0, y0 - 8),
                  min(img.width, x1 + 9), min(img.height, y1 + 9))
                 ).save(os.path.join(crop_dir, f'row_{i}.png'))

    return {
        '사진': os.path.basename(path),
        '채택': {
            'method': 'template@auto',   # 지금 실제로 쓰는 방법(get_text_mask)
            'line_count': line_count,
            'rows': [[r['번호'], r['값'], r['X'], r['Y'], bool(r['확인필요'])]
                     for r in rows],
        },
        # 목록 형태로 두는 이유: 나중에 다른 엔진을 더해도 칸 구조를 안 바꿔도 됨
        'attempts': [{'method': a['method'], 'line_count': a['line_count'],
                      'texts': a['texts'],
                      'rows': [[r['번호'], r['값'], r['X'], r['Y']] for r in a['rows']]}
                     for a in attempts],
    }


def read_measured(files, templates, plan, progress_cb=None, run_dir=None):
    """process_folder의 1단계(OCR 읽기)만 떼어낸 것.

    ⚠️ 왜 떼어냈나(2026-08-07): 웹은 처리가 백그라운드 스레드에서 도는데,
    스레드는 브라우저의 폼 제출(사람 입력)을 기다릴 수 없음. 그래서 웹에서는
    여기까지만 하고 스레드를 끝낸 뒤(진행률에 '입력 대기' 상태만 남기고),
    사람이 폼을 제출하면 그 요청 안에서 merge_manual_entries + finish_processing을
    이어서 부름. 데스크탑은 이 셋을 process_folder가 그대로 이어 붙여서 씀 -
    tkinter 창이 같은 프로세스 안에서 동기적으로 답을 기다릴 수 있기 때문.

    반환: (measured, log, evidence_records) - 셋 다 JSON으로 그대로 저장 가능한
    기본 자료형(문자열·숫자·불린·리스트·딕셔너리)만 담고 있음. 웹이 스레드 종료
    전에 파일로 남겼다가, 나중 요청에서 그대로 복원해 이어서 쓸 수 있어야 하므로."""
    log = []

    # 1) 사진마다 OCR로 읽고, 측정값이 있는 것만 남김
    #    종류는 "읽은 값 개수"가 아니라 "실제 측정줄 개수"로 판별함. 계측기가 두
    #    측정줄을 겹쳐 그리면 그 줄은 글자가 뭉개져 못 읽는데, 읽은 개수로 종류를
    #    정하면 4줄짜리 overlay 사진이 2줄짜리 Line&Space로 둔갑해서 엉뚱한
    #    Target에 붙어버림. 줄 개수는 박스 폭으로 따로 셀 수 있어 이걸 씀.
    measured = []
    evidence_records = []
    for i, f in enumerate(files):
        name = os.path.basename(f)
        rows, line_count = read_image_ex(f, templates)
        if line_count == 0:
            log.append(f'{name} → 측정 전 사진 (버림)')
            if progress_cb:
                progress_cb(i + 1, len(files))
            continue

        category = KIND_LABEL_BY_COUNT.get(line_count)

        # 측정줄이 1개인 사진은 CD일 수도, "한쪽만 있는 L/S Item"일 수도 있음
        # (Line만 있는 Item은 값이 1개만 찍혀서 CD 사진과 생김새가 같음).
        # 줄 개수로는 못 가리므로 값으로 정한다. 2026-08-05.
        if category == 'CD' and rows:
            resolved = match_single_target(rows[0]['값'], plan)
            if resolved:
                category = resolved[0]

        kind = display_kind(category, plan) if category else f'알수없음({line_count}줄)'

        missing = line_count - len(rows)
        if missing > 0:
            # 값이 뭉개진 사진은 남은 값도 믿기 어려우니 전부 확인 대상으로 표시
            for r in rows:
                r['확인필요'] = True
                r['확인사유'] = (r.get('확인사유')
                               or f'같은 사진에서 {missing}줄을 못 읽음')
            log.append(f'{name} → {kind}, 값 {len(rows)}개 읽음 (총 {line_count}줄)')
            log.append(f'  {UNREAD_MARK} {name}: 측정값 {line_count}줄 중 {missing}줄을 읽지 못했습니다 '
                        '(측정값 두 줄이 겹쳐 그려졌거나 글자가 뭉개진 경우) — '
                        '사진을 열어 값을 직접 확인해 주세요.')
        else:
            log.append(f'{name} → {kind}, 값 {len(rows)}개 읽음')

        # 판독에 문제가 있는 사진만 근거를 남김. 전부 남기면 ①밝기 기준마다
        # 사진을 다시 읽어야 해서 느려지고(77장 11초가 몇 배로) ②정작 봐야 할
        # 사진이 묻힘. 전수 조사가 필요하면 별도 분석 스크립트로 할 것.
        if run_dir and (missing > 0 or any(r['확인필요'] for r in rows)):
            evidence_records.append(
                collect_ocr_evidence(run_dir, f, rows, line_count, templates))

        measured.append((f, name, category, kind, rows, line_count))
        if progress_cb:
            progress_cb(i + 1, len(files))

    return measured, log, evidence_records


def merge_manual_entries(measured, manual_by_path, plan, log):
    """process_folder의 1.5단계 - 못 읽은 값을 사람이 직접 입력한 결과를 합침.

    Target 매칭(2단계) 전에 해야 손으로 넣은 값도 OCR 값과 완전히 같은 검증
    (Target 매칭·이상치 검사)을 통과함. manual_by_path가 비어있으면 원본 그대로
    반환(변경 없음). log는 호출자 것을 그대로 이어서 씀(같은 리스트를 반환)."""
    if not manual_by_path:
        return measured, log
    merged = []
    for f, name, category, kind, rows, line_count in measured:
        extra = manual_by_path.get(f)
        if extra:
            before = len(rows)
            rows = rows + extra
            # 완전히 채워졌으면 "같은 사진에서 N줄을 못 읽음"이라는
            # 사진 전체 의심 사유는 더 이상 맞지 않으므로 지움.
            # (그 행에 다른 개별 사유가 있었다면 or 때문에 안 덮였으므로 그대로 남음)
            if len(rows) >= line_count:
                stale = f'같은 사진에서 {line_count - before}줄을 못 읽음'
                for r in rows[:before]:
                    if r.get('확인사유') == stale:
                        r['확인필요'] = False
                        r['확인사유'] = ''
            # CD 사진 1줄짜리는 종류를 값으로 정하는데(2026-08-05), 원래
            # 읽은 값이 없어 못 정했던 경우 손으로 넣은 값으로 다시 정함.
            if category == 'CD' and rows:
                resolved = match_single_target(rows[0]['값'], plan)
                if resolved:
                    category = resolved[0]
                    kind = display_kind(category, plan)
            log.append(f'{name} → 수동입력 {len(extra)}개 추가 '
                       f'(이제 {len(rows)}/{line_count})')
        merged.append((f, name, category, kind, rows, line_count))
    return merged, log


def finish_processing(measured, plan, log, evidence_records, run_dir=None):
    """process_folder의 2)+3)단계 - Target 매칭부터 끝까지.
       (표 데이터, Overlay 요약, 로그 목록)을 반환."""
    all_rows = []

    # 2) Target 매칭 (측정값마다 제일 가까운 Item을 찾아 편차·Tolerance 계산)
    for path, name, category, kind, rows, _line_count in measured:
        for r in rows:
            r['방향'] = ''  # overlay가 아니면 방향 없음

        # 값이 1개짜리 사진(CD 또는 한쪽만 있는 L/S Item)은 같은 경로로 매칭.
        # category는 위에서 이미 값으로 정해졌으므로 여기서는 항목 이름만 붙임.
        if category == 'CD' or (category == 'L/S' and len(rows) == 1):
            for r in rows:
                match = match_single_target(r['값'], plan)
                if match:
                    _, r['항목'], r['Target'], r['편차'], r['Tolerance'], ambiguous = match
                    if ambiguous:
                        # CD Target과 L/S Target이 서로 가까워 어느 쪽인지 확신 불가
                        r['확인필요'] = True
                        r['확인사유'] = (r.get('확인사유')
                                       or 'CD/L&S 어느 쪽인지 확신 불가')
                        log.append(f'  [확인] {name}: 값 {r["값"]}이(가) CD와 Line&Space '
                                   '양쪽 규격에 모두 들어와 종류를 확정할 수 없습니다 — '
                                   '사진을 확인해 주세요.')
                else:
                    r['항목'], r['Target'], r['편차'], r['Tolerance'] = '', '', '', ''
        elif category == 'L/S':
            vals = [r['값'] for r in rows]
            match = match_ls_targets(vals[0], vals[1], plan) if len(vals) == 2 else None
            if match:
                item_name, line_val, t_line, d_line, tol_line, space_val, t_space, d_space, tol_space = match
                line_row = space_row = None
                for r in rows:
                    if r['값'] == line_val:
                        r['항목'], r['Target'], r['편차'], r['Tolerance'] = f'{item_name}-Line', t_line, d_line, tol_line
                        line_row = r
                    else:
                        r['항목'], r['Target'], r['편차'], r['Tolerance'] = f'{item_name}-Space', t_space, d_space, tol_space
                        space_row = r

                # 2차 검증: Line/Space Target이 같아 크기로만 배정한 경우에 한해,
                # 크로스헤어 밝기가 "큰 값=Line" 배정과 맞는지 확인 (과제 4)
                if line_row is not None and space_row is not None and t_line == t_space:
                    agrees = verify_ls_brightness(path, line_row, space_row)
                    if agrees is False:
                        for rr in (line_row, space_row):
                            rr['확인필요'] = True
                            rr['확인사유'] = (rr.get('확인사유')
                                            or 'Line/Space 배정이 밝기 검증과 다름')
                        log.append(f'  [확인] {name}: Line/Space 배정(크기 기준)이 크로스헤어 '
                                    '밝기 검증과 다릅니다 — 사진을 확인해 주세요.')
            else:
                for r in rows:
                    r['항목'], r['Target'], r['편차'], r['Tolerance'] = '', '', '', ''
        elif category == 'overlay':
            assign_directions(rows)  # 화면 위치로 상/하/좌/우 판별 (과제 2)
            for r in rows:
                r['항목'], r['Target'], r['편차'], r['Tolerance'] = '', '', '', ''
        else:  # 알수없음 등 -> Target 비교 대상 아님
            for r in rows:
                r['항목'], r['Target'], r['편차'], r['Tolerance'] = '', '', '', ''

        # USL/LSL 및 스펙이탈 여부
        for r in rows:
            if r['Tolerance'] not in ('', None):
                r['USL'] = round(r['Target'] + r['Tolerance'], 4)
                r['LSL'] = round(r['Target'] - r['Tolerance'], 4)
                r['스펙이탈'] = '예' if not (r['LSL'] <= r['값'] <= r['USL']) else ''
            else:
                r['USL'], r['LSL'], r['스펙이탈'] = '', '', ''

    # 3) 사진을 Point 단위로 묶어 번호를 매김.
    #    한 Point의 사진 장수는 공정마다 다름 — 예전엔 무조건 3장씩(overlay/CD/L-S
    #    한 장씩) 묶고 "종류가 셋 다 달라야 정상"으로 봤는데, CD Item이 2개 이상인
    #    공정(예: Bump의 25CD/28CD Item 2개)에서는 한 Point에 '종류=CD' 사진이
    #    2장 나오는 게 정상인데도 종류가 겹친다고 매번 오탐이 났음(2026-07-30,
    #    실제 Bump 실행에서 13개 Point 전부 걸림). 그래서 "종류가 다 달라야 함"이
    #    아니라 "계획에 입력한 개수(CD Item 수 + L/S Item 수 + Overlay 유무)만큼
    #    맞게 있는가"로 판단 기준을 바꿈.
    expected_counts = Counter({
        'CD': len(plan.cd_items),
        'L/S': len(plan.ls_items),
        'overlay': 1 if plan.overlay_points else 0,
    })
    expected_counts = +expected_counts  # 0개인 항목은 비교에서 자동으로 빠짐
    group_size = sum(expected_counts.values()) or 3  # 계획이 통째로 비어있으면 예전 기본값

    overlay_summaries = []
    point_no = 0
    for i in range(0, len(measured), group_size):
        group = measured[i:i + group_size]
        point_no += 1
        actual_counts = Counter(category for _, _, category, _, _, _ in group)
        if len(group) != group_size or actual_counts != expected_counts:
            log.append(f'  [경고] Point {point_no}: 한 세트(overlay/CD/L-S)가 맞지 않습니다 '
                        f'({[k for _, _, _, k, _, _ in group]}) → 사진을 확인해 주세요.')

        for path, name, category, kind, rows, _line_count in group:
            if category == 'overlay':
                metrics = compute_overlay_metrics(rows)
                if metrics:
                    overlay_summaries.append({'Point': point_no, '파일명': name, **metrics})
                else:
                    log.append(f'  [경고] {name}: 상/하/좌/우 4방향이 모두 확인되지 않아 '
                                'Overlay 지표를 계산하지 못했습니다.')

            for r in rows:
                all_rows.append({
                    'Point': point_no,
                    '파일명': name,
                    '종류': kind,
                    '항목': r['항목'],
                    '방향': r['방향'],
                    '번호': r['번호'],
                    '측정값(um)': r['값'],
                    'Target': r['Target'],
                    '편차': r['편차'],
                    'Tolerance': r['Tolerance'],
                    'USL': r['USL'],
                    'LSL': r['LSL'],
                    '스펙이탈': r['스펙이탈'],
                    'X(um)': r['X'],
                    'Y(um)': r['Y'],
                    '확인필요': '예' if r['확인필요'] else '',
                    '확인사유': r.get('확인사유', ''),
                    '입력방법': r.get('입력방법', '자동'),
                    '신뢰도': r['신뢰도'],
                })

    # 판독 이상치 경보 — 같은 항목의 동료 값과 비교해야 하므로 모든 행이
    # 갖춰진 뒤에 본다. 값은 바꾸지 않고 확인필요 표시만 붙임.
    flagged = set()
    for r, reason in find_ocr_outliers(all_rows):
        if id(r) in flagged:
            continue          # 두 검사에 다 걸린 행은 먼저 나온 사유만 남김
        flagged.add(id(r))
        r['확인필요'] = '예'
        r['확인사유'] = r.get('확인사유') or reason
        log.append(f'  [확인] {r["파일명"]} 번호{r["번호"]}: {reason}')

    # 판독 근거(글)는 실행 하나당 한 파일로 모아서 남김
    if run_dir and evidence_records:
        out = os.path.join(run_dir, OCR_EVIDENCE_DIR, OCR_EVIDENCE_JSON)
        os.makedirs(os.path.dirname(out), exist_ok=True)
        with open(out, 'w', encoding='utf-8') as f:
            json.dump(evidence_records, f, ensure_ascii=False, indent=1)

    return all_rows, overlay_summaries, log


def process_folder(files, templates, plan, progress_cb=None, run_dir=None,
                   manual_input_cb=None):
    """사진 목록(시간순 정렬됨)을 읽어서 (표 데이터, Overlay 요약, 로그 목록)을 반환.
       progress_cb(읽은 장수, 전체 장수)가 있으면 사진 한 장 읽을 때마다 호출함
       (웹 버전에서 "N/전체 처리 중" 진행률 표시에 씀 - 데스크탑은 안 넘기므로
       그대로 동작에 영향 없음).

       manual_input_cb가 있으면 1단계(읽기)와 2단계(Target 매칭) 사이에 불러서
       못 읽은 값을 사람이 직접 넣을 기회를 줌(manual_entry.py). 여기서 끼워
       넣는 이유는 손으로 넣은 값도 OCR 값과 완전히 같은 경로(Target 매칭·
       이상치 검사)를 지나가야 하기 때문. 안 넘기면 지금까지와 동작이 완전히 같음.

       ⚠️ 이 함수는 read_measured + merge_manual_entries + finish_processing을
       한 번에 이어 붙인 것뿐(데스크탑 전용 - 같은 프로세스 안에서 tkinter 창이
       동기적으로 답을 기다릴 수 있어서 가능함). 웹은 스레드가 사람 입력을
       기다릴 수 없어서 이 셋을 요청 두 번(처리 시작 / 폼 제출)에 나눠 부름 -
       webapp/app.py의 run_processing·view_manual_input 참고(2026-08-07)."""
    measured, log, evidence_records = read_measured(
        files, templates, plan, progress_cb, run_dir)
    if manual_input_cb:
        manual_by_path = manual_input_cb(measured)
        measured, log = merge_manual_entries(measured, manual_by_path, plan, log)
    return finish_processing(measured, plan, log, evidence_records, run_dir)


# ------------------------------------------------------------
#  엑셀 꾸미기: 스펙이탈 셀 색칠 + Cpk 막대그래프 + Item별 트렌드 차트
# ------------------------------------------------------------
# 이 엑셀은 "보고 끝내는 문서"가 아니라 사내 양식으로 옮겨 붙이는 원본이다
# (사용자 확인: 측정결과·Cpk요약 두 시트를 통째로 복사해서 씀). 그래서 꾸미기의
# 기준이 "예쁜가"가 아니라 "복사해서 붙였을 때 상대 양식을 망치지 않는가"다.
#   - 색은 조건부 서식이 아니라 셀에 직접 칠한다. 조건부 서식은 붙여넣을 때
#     규칙째 따라가서 상대 양식에 남의 규칙을 심어 버린다(2026-08-21 결정).
#   - 머리글은 굵게만. 배경색을 넣으면 머리글째 복사할 때 회색이 따라붙는다.
#   - 셀 병합·제목행·로고는 넣지 않는다. 복사 범위가 어긋난다.
SPEC_RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
SPEC_YELLOW = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
SPEC_GREEN = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

TAB_COPY_TARGET = '1C3F5F'   # 사내 양식으로 복사해 가는 두 시트
TAB_REFERENCE = 'A6ADB4'     # 참고용 시트

# 표 모양을 갖춘 시트만 머리글 고정·자동 필터 대상. 트렌드차트 시트는 Item마다
# 표가 따로 있는 다단 배치라 여기에 넣으면 안 된다(필터가 첫 표만 걸림).
TABLE_SHEETS = ('측정결과', '검증요약', 'Overlay요약', 'Cpk요약')


def _cell_width(value):
    """엑셀 열 너비 계산용 글자 폭. 한글·전각 문자는 두 칸으로 센다."""
    text = '' if value is None else str(value)
    return sum(2 if unicodedata.east_asian_width(c) in ('W', 'F') else 1 for c in text)


def _decimal_places(value):
    """이 값이 실제로 쓰는 소수 자리수. 숫자가 아니면 None."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return None
    text = repr(float(value))
    if 'e' in text or 'E' in text:
        return 6
    return len(text.split('.')[1].rstrip('0')) if '.' in text else 0


def _number_format(max_places):
    """자리수를 코드에 박지 않는 이유: 공정마다 단위가 달라서(RDL은 100um대,
    L&S는 5um대) 하나로 정하면 한쪽이 반올림된 것처럼 보인다. 대신 열에 실제로
    들어 있는 값의 최대 자리수를 세서 서식을 만든다 — 한 열 안에서는 자리수가
    늘 같으므로 이렇게 하면 값이 깎이는 일이 없다."""
    return '0' if max_places <= 0 else '0.' + '0' * min(max_places, 6)


def polish_workbook(writer):
    """붙여넣기 안전한 최소 정비 — 열 너비, 머리글 굵게+틀 고정+자동 필터,
    열별 숫자 표시 서식, 탭 색."""
    for name, ws in writer.sheets.items():
        ws.sheet_properties.tabColor = (
            TAB_COPY_TARGET if name in ('측정결과', 'Cpk요약') else TAB_REFERENCE)
        if name not in TABLE_SHEETS or ws.max_row < 1:
            continue

        for cell in ws[1]:
            if cell.value is not None:
                cell.font = Font(bold=True)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        for column in ws.iter_cols(min_row=1, max_row=ws.max_row):
            places = [_decimal_places(c.value) for c in column[1:]]
            places = [x for x in places if x is not None]
            fmt = _number_format(max(places)) if places else None
            width = _cell_width(column[0].value)
            for cell in column[1:]:
                if fmt and _decimal_places(cell.value) is not None:
                    cell.number_format = fmt
                    width = max(width, len(f'{cell.value:.{max(places)}f}'))
                else:
                    width = max(width, _cell_width(cell.value))
            # 8은 날짜·짧은 머리글이 눌리지 않는 하한, 40은 '확인사유'처럼 문장이
            # 들어가는 열이 화면을 다 먹지 않게 하는 상한.
            ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 8), 40)


def style_result_sheet(writer, df):
    ws = writer.sheets['측정결과']
    spec_col = df.columns.get_loc('스펙이탈') + 1
    for row in range(2, len(df) + 2):
        cell = ws.cell(row=row, column=spec_col)
        if cell.value == '예':
            cell.fill = SPEC_RED


def add_cpk_sheet(writer, stats_rows):
    if not stats_rows:
        return
    # 앞쪽 열 순서는 기존에 쓰던 양식 그대로 두고, 새로 추가된 지표(3σ/CDU%/판정)만
    # 뒤에 붙임 — 기존 양식을 보던 사람이 열 위치를 다시 익히지 않아도 되게.
    stats_df = pd.DataFrame(stats_rows)[
        ['표시명', 'N', 'Target', 'Tolerance', 'LSL', 'USL', 'Cp', 'Cpk',
         'Avg', 'Std', 'Max', 'Min', 'Range', '3σ', 'CDU%', '판정']
    ].rename(columns={'표시명': '항목', 'Tolerance': 'Tol.', 'Cp': 'CP', 'Cpk': 'CPK',
                       'Std': 'Std (1σ)'})
    stats_df.to_excel(writer, sheet_name='Cpk요약', index=False)

    ws = writer.sheets['Cpk요약']
    n = len(stats_df) + 1

    # CPK 등급 색칠: 1.0 미만 빨강, 1.33까지 노랑, 그보다 크면 초록.
    # ⚠️ 딱 1.33인 값은 노랑이다 — 전에 쓰던 조건부 서식에서 '1.0~1.33 사이' 규칙이
    # '1.33 이상' 규칙보다 우선순위가 높아서 그렇게 보였고, 색칠 방식만 바꾸는
    # 김에 보이는 색까지 바뀌면 안 되므로 그대로 뒀다.
    cpk_col = stats_df.columns.get_loc('CPK') + 1
    verdict_col = stats_df.columns.get_loc('판정') + 1
    for row in range(2, n + 1):
        cell = ws.cell(row=row, column=cpk_col)
        if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
            if cell.value < 1.0:
                cell.fill = SPEC_RED
            elif cell.value <= 1.33:
                cell.fill = SPEC_YELLOW
            else:
                cell.fill = SPEC_GREEN

        # 판정 색칠: FAIL은 빨강, PASS는 초록 (측정결과 시트의 스펙이탈 색과 같은 계열)
        verdict = ws.cell(row=row, column=verdict_col)
        if verdict.value == 'FAIL':
            verdict.fill = SPEC_RED
        elif verdict.value == 'PASS':
            verdict.fill = SPEC_GREEN

    # CPK 막대그래프
    chart = BarChart()
    chart.title = 'Item별 CPK'
    chart.y_axis.title = 'CPK'
    data = Reference(ws, min_col=stats_df.columns.get_loc('CPK') + 1, min_row=1, max_row=n)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f'A{n + 3}')


def add_trend_sheet(writer, df, plan):
    """Item(항목)별로 Point 순서에 따른 측정값 추이를 선그래프로."""
    items = [x for x in df['항목'].unique() if x]
    if not items:
        return
    sheet_name = '트렌드차트'
    ws_data_row = 1
    startcol = 0
    workbook = writer.book
    ws = workbook.create_sheet(sheet_name)

    for item in items:
        sub = df[df['항목'] == item][['Point', '측정값(um)', 'Target', 'USL', 'LSL']]
        if sub.empty:
            continue
        display = display_item_name(item, sub['Target'].iloc[0], plan)
        header_row = ws_data_row
        ws.cell(row=header_row, column=startcol + 1, value=f'[{display}]')
        for j, colname in enumerate(sub.columns):
            ws.cell(row=header_row + 1, column=startcol + 1 + j, value=colname)
        for i, (_, row) in enumerate(sub.iterrows()):
            for j, colname in enumerate(sub.columns):
                ws.cell(row=header_row + 2 + i, column=startcol + 1 + j, value=row[colname])

        first_data_row = header_row + 1
        last_data_row = header_row + 1 + len(sub)

        chart = LineChart()
        chart.title = f'{display} 추이'
        chart.y_axis.title = '측정값(um)'
        chart.x_axis.title = 'Point'
        data = Reference(ws, min_col=startcol + 2, max_col=startcol + 5,
                          min_row=first_data_row, max_row=last_data_row)
        cats = Reference(ws, min_col=startcol + 1, min_row=first_data_row + 1, max_row=last_data_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f'A{header_row + len(sub) + 4}')

        ws_data_row = header_row + len(sub) + 22  # 다음 Item 표 시작 위치(차트 자리 확보)


# ------------------------------------------------------------
#  [자연어 정렬] 파일명 속 숫자를 "숫자"로 보고 정렬하기 위한 도우미 함수
# ------------------------------------------------------------
def natural_sort_key(file_path):
    # 문제 상황: 파이썬 기본 정렬(sorted())은 파일명을 글자 하나씩 비교하는
    # "문자열 정렬"이라서, 숫자가 포함된 파일명을 사람이 생각하는 순서와
    # 다르게 정렬해버림. 예를 들어 "10 (43).bmp"와 "2 (46).bmp"를 비교하면
    # 맨 앞 글자 '1'이 '2'보다 작으므로 "10..."이 "2..."보다 앞에 온다고
    # 판단함 → 실제로는 2, 3, ... 9 다음에 10이 와야 하는데 순서가 깨짐.
    # (실제로 LS Sample 폴더의 "1 (45).bmp", "10 (43).bmp" 같은 파일명에서
    #  이 문제가 발생함. 반면 CaptImg20260619092246.bmp 처럼 숫자 자릿수가
    #  항상 같은 파일명은 문자열 정렬로도 우연히 순서가 맞았음.)
    #
    # 해결 방법("자연어 정렬", natural sort): 파일명을 숫자 부분과
    # 숫자가 아닌 부분으로 잘게 쪼갠 뒤, 숫자 부분은 진짜 숫자(int)로 바꿔서
    # 비교함. 그러면 "10"은 10이라는 숫자로, "2"는 2라는 숫자로 비교되니까
    # 10이 2보다 크다고 정확히 판단하게 됨.
    #
    # 예) "10 (43).bmp" → ['', 10, ' (', 43, ').bmp']  (숫자는 int로 변환됨)
    #     "2 (46).bmp"  → ['', 2, ' (', 46, ').bmp']
    #     위 두 리스트를 비교하면 10 > 2 이므로 "2 (46).bmp"가 먼저 옴 — 정상.
    file_name = os.path.basename(file_path)
    pieces = re.split(r'(\d+)', file_name)  # 숫자 앞뒤로 문자열을 나눔
    return [int(piece) if piece.isdigit() else piece for piece in pieces]


def _launch(path):
    """파일 하나를 기본 프로그램으로 엶. 성공하면 True."""
    # HTML은 브라우저가 경로를 "주소"로 받아들이기 때문에, 폴더 이름에 '#'나
    # 공백이 있으면 그 뒤가 잘려버릴 수 있음 (예: "#1 Lot" 같은 실제 작업 폴더).
    # 그래서 미리 안전한 주소 형태(file:///...%231%20Lot/...)로 바꿔서 넘김.
    if path.lower().endswith(('.html', '.htm')):
        try:
            if webbrowser.open(Path(path).as_uri()):
                print(f'[열기·브라우저] {path}')
                return True
        except Exception as e:
            print(f'[알림] 브라우저로 열기 실패 ({e}): {path}')

    try:
        os.startfile(path)   # 윈도우 전용, 확장자에 연결된 기본 프로그램
        print(f'[열기] {path}')
        return True
    except Exception as e:
        print(f'[알림] os.startfile 실패 ({e}) → cmd start로 다시 시도합니다: {path}')

    # 파일 연결이 없거나 startfile이 막힌 환경 대비 마지막 시도
    try:
        subprocess.Popen(['cmd', '/c', 'start', '', path], shell=False)
        print(f'[열기·2차] {path}')
        return True
    except Exception as e:
        print(f'[알림] 파일을 자동으로 열지 못했습니다 — 직접 열어 주세요: {path} ({e})')
        return False


def open_saved_files(paths):
    """저장된 결과 파일을 각 확장자에 연결된 기본 프로그램으로 열어줌
       (.html → 브라우저, .xlsx → 엑셀).
       파일은 이미 저장이 끝난 상태이므로, 여는 데 실패해도 프로그램을
       멈추지 않고 안내만 함."""
    failed = []
    for p in paths:
        # askdirectory()가 돌려주는 경로는 슬래시(/)가 섞여 있어서 윈도우 표준
        # 형태로 정리해둠. 절대경로로 바꿔야 실행 위치와 무관하게 열림.
        p = os.path.normpath(os.path.abspath(p))
        if not os.path.exists(p):
            print(f'[알림] 열려던 파일이 없습니다: {p}')
            failed.append(p)
            continue

        if not _launch(p):
            failed.append(p)

    # 실행 요청이 윈도우 셸에 전달될 시간을 잠깐 줌. 파이썬이 곧바로 종료되면
    # 프로그램이 뜨기 전에 끊기는 경우가 있음.
    time.sleep(1.5)

    # 콘솔 출력은 못 보고 지나치기 쉬우므로, 실패는 창으로 알려줌
    if failed:
        messagebox.showwarning(
            '자동 열기 실패',
            '아래 파일을 자동으로 열지 못했습니다. 폴더에서 직접 열어 주세요:\n\n'
            + '\n'.join(failed))


def save_run_log(folder, plan, files, log, df, summary_df, out_of_spec, flagged, unread,
                 operator='', elapsed_sec=None, category=''):
    """실행 결과를 두 곳에 남김.
       (1) 상세 로그 — 사진 폴더 안에 실행마다 새 파일. 사진·엑셀·HTML과
           한 폴더에 있으니 나중에 로그의 파일명으로 바로 원본 사진을 찾을 수 있음.
       (2) 요약 통계 — 프로그램 폴더의 실행이력.csv에 한 줄씩 누적. 여러 Lot(폴더)를
           오가며 실행해도 한곳에 모여 있어야 추세(예: 확인필요 비율이 늘고 있는지)를
           비교해볼 수 있어서 사진 폴더가 아니라 여기에 둠."""
    now = datetime.datetime.now()
    timestamp = now.strftime('%Y%m%d_%H%M%S')
    mismatched_points = sum(1 for line in log if '[경고]' in line and '한 세트' in line)
    mismatched_items = int((summary_df['일치'] == '아니오').sum()) if not summary_df.empty else 0

    # (1) 상세 로그
    log_path = os.path.join(folder, f'측정로그_{timestamp}.txt')
    header = [
        f'실행 시각: {now.strftime("%Y-%m-%d %H:%M:%S")}',
        f'공정: {plan.process} ({plan.cd_label})',
        f'고객/자재/Layer: {plan.customer or "-"} / {plan.material or "-"} / {plan.layer or "-"}',
        f'폴더: {folder}',
        f'입력 사진: {len(files)}장',
    ]
    # 웹에서 돌린 경우에만 붙음(데스크탑은 로그인 개념이 없어 비어 있음).
    # 없는 값을 '-'로라도 넣지 않는 이유: 로그를 눈으로 읽을 때 의미 없는 줄이
    # 늘어나기만 해서.
    if operator:
        header.append(f'실행자: {operator}')
    if elapsed_sec is not None:
        header.append(f'소요 시간: {elapsed_sec:.1f}초')
    header += ['', '--- 사진별 처리 내역 ---']
    footer = [
        '',
        '--- 요약 ---',
        f'총 측정값: {len(df)}개 (Point {df["Point"].nunique() if "Point" in df.columns else 0}개)',
        f'확인필요: {flagged}개',
        f'스펙이탈: {out_of_spec}개',
        f'겹침(측정줄 일부 못 읽음): {unread}장',
        f'세트 불일치 Point: {mismatched_points}개',
        f'포인트 수 불일치 Item: {mismatched_items}개',
    ]
    try:
        with open(log_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(header + log + footer) + '\n')
    except Exception as e:
        print(f'[알림] 상세 로그 저장 실패: {e}')
        log_path = None

    # (2) 요약 통계 누적
    _migrate_run_history_header()
    is_new = not os.path.exists(RUN_HISTORY_CSV)
    row = [
        now.strftime('%Y-%m-%d %H:%M:%S'), folder, plan.process,
        plan.customer, plan.material, plan.layer,
        len(files), len(files) - sum(1 for line in log if '측정 전 사진' in line),
        len(df), df['Point'].nunique() if 'Point' in df.columns else 0,
        flagged, out_of_spec, unread, mismatched_points, mismatched_items,
        operator, '' if elapsed_sec is None else round(elapsed_sec, 1), category,
    ]
    try:
        with open(RUN_HISTORY_CSV, 'a', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            if is_new:
                writer.writerow(RUN_HISTORY_HEADER)
            writer.writerow(row)
    except Exception as e:
        print(f'[알림] 실행이력.csv 기록 실패: {e}')

    return log_path


# ------------------------------------------------------------
#  [메인]
# ------------------------------------------------------------
def main():
    result = collect_plan_and_folder()
    if result is None:
        return  # 사용자가 취소함
    plan, folder = result

    # 장비 설정/버전에 따라 저장 형식이 달라서(.bmp / .png / .jpg) 모두 지원함.
    # OCR은 PIL로 픽셀만 읽으므로 형식과 무관하게 같은 방식으로 동작함.
    files = []
    for ext in IMAGE_EXTENSIONS:
        files += glob.glob(os.path.join(folder, f'*{ext}'))
        files += glob.glob(os.path.join(folder, f'*{ext.upper()}'))
    # 예전엔 sorted(set(files))로 문자열 그대로 정렬했는데, 파일명에 숫자가
    # 들어간 경우(예: "LS Sample"의 "1 (45).bmp") 순서가 꼬이는 문제가 있었음.
    # natural_sort_key로 숫자를 진짜 숫자로 인식해서 정렬하도록 고침.
    files = sorted(set(files), key=natural_sort_key)

    if not files:
        messagebox.showerror(
            '오류', '이 폴더에 사진이 없습니다.\n'
            f'({" / ".join(IMAGE_EXTENSIONS)} 형식을 읽을 수 있습니다)')
        return

    # jpg는 손실 압축이라 글자 가장자리가 미세하게 뭉개짐. 이 프로그램은 글자
    # 모양을 템플릿과 비교해 읽으므로 6을 5로, 0을 3으로 착각하는 일이 생김
    # (실제 사진 166장을 jpg로 바꿔 대조하니 21장에서 판독이 달라졌음).
    # 다행히 잘못 읽은 값은 전부 "확인필요"로 걸러졌지만, 사람이 그만큼 더
    # 확인해야 하므로 미리 알려줌.
    lossy = [f for f in files if f.lower().endswith(('.jpg', '.jpeg'))]
    if lossy:
        messagebox.showwarning(
            'JPG 사진 안내',
            f'JPG 사진이 {len(lossy)}장 있습니다.\n\n'
            'JPG는 압축 과정에서 글자가 살짝 뭉개져서, 숫자를 잘못 읽을 수 있습니다 '
            '(잘못 읽은 값은 "확인필요"로 표시되니 그 값들은 사진을 직접 확인해주세요).\n\n'
            '가능하면 계측기에서 BMP나 PNG로 저장하면 이런 문제가 없습니다.')

    templates = load_templates()
    # tkinter 창을 여기서만 쓰므로 import도 여기서만 함 - 이 파일은 웹에서도
    # import되는데, manual_entry는 웹에서 아직 안 쓰고(2026-08-07 기준 데스크탑
    # 전용 검증 단계) 화면 라이브러리를 웹 쪽까지 끌고 갈 이유가 없음.
    from manual_entry import ask_manual_values
    # 근거는 사진과 같은 폴더 안에 남김(결과 엑셀·로그가 가는 곳과 동일)
    all_data, overlay_summaries, log = process_folder(
        files, templates, plan, run_dir=folder, manual_input_cb=ask_manual_values)

    print('\n'.join(log))

    if not all_data:
        messagebox.showerror('오류', '어떤 사진에서도 측정값을 읽지 못했습니다.')
        return

    df = pd.DataFrame(all_data)[
        # ⚠️ '스펙이탈'과 '확인필요'는 뜻이 다름 — 스펙이탈은 공정 결과가 규격
        # 밖이라는 것이고, 확인필요는 판독을 믿기 어렵다는 뜻. '확인사유'에
        # 무엇 때문인지 적힘(2026-08-05 추가). '입력방법'은 못 읽은 값을 사람이
        # 직접 넣었는지 구분(2026-08-07 추가, manual_entry.py) — 계측 데이터라
        # 손으로 넣은 값이 기계가 읽은 값과 안 섞이고 추적돼야 함.
        ['Point', '파일명', '종류', '항목', '방향', '번호', '측정값(um)', 'Target', '편차',
         'Tolerance', 'USL', 'LSL', '스펙이탈', 'X(um)', 'Y(um)', '확인필요', '확인사유',
         '입력방법', '신뢰도']
    ]
    summary_df = pd.DataFrame(summarize_matching(all_data, plan))
    stats_rows = compute_stats(all_data, plan)
    overlay_df = pd.DataFrame(overlay_summaries)[
        ['Point', '파일명', '상', '하', '좌', '우', 'Overlay_X', 'Overlay_Y', 'Overlay']
    ] if overlay_summaries else pd.DataFrame()

    save_path = os.path.join(folder, '측정결과.xlsx')
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='측정결과', index=False)
        style_result_sheet(writer, df)
        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name='검증요약', index=False)
        if not overlay_df.empty:
            overlay_df.to_excel(writer, sheet_name='Overlay요약', index=False)
        add_cpk_sheet(writer, stats_rows)
        add_trend_sheet(writer, df, plan)
        polish_workbook(writer)   # 시트가 다 만들어진 뒤에 한 번에 정비

    item_breakdown = compute_item_breakdown(all_data, plan, unread_by_file(log))
    report_path = os.path.join(folder, '측정결과.html')
    build_report(df, overlay_df, stats_rows, plan, report_path, item_breakdown=item_breakdown)

    flagged = (df['확인필요'] == '예').sum()
    out_of_spec = (df['스펙이탈'] == '예').sum()
    msg = f'완료! 파일을 저장했습니다:\n{save_path}\n{report_path}'
    if flagged:
        msg += f'\n\n※ 확인이 필요한 값 {flagged}개가 있습니다 ("확인필요" 열 참고)'
    if out_of_spec:
        msg += f'\n※ 스펙(USL/LSL) 벗어난 값 {out_of_spec}개가 있습니다 ("스펙이탈" 열 참고)'
    if not summary_df.empty and (summary_df['일치'] == '아니오').any():
        msg += '\n※ 예상 포인트 수와 실제 매칭 개수가 다른 항목이 있습니다 ("검증요약" 시트 참고)'
    unread = sum(1 for line in log if UNREAD_MARK in line)
    if unread:
        msg += (f'\n※ 사진 {unread}장에서 측정값 일부를 읽지 못했습니다 (측정값 두 줄이 겹쳐 '
                '그려졌거나 글자가 뭉개진 경우). 해당 사진은 직접 확인해주세요 '
                '(자세한 파일명은 실행 창에 표시됨)')

    log_path = save_run_log(folder, plan, files, log, df, summary_df, out_of_spec, flagged, unread)
    if log_path:
        msg += f'\n\n로그: {log_path}\n(실행 이력 요약은 프로그램 폴더의 실행이력.csv에 누적됨)'

    msg += '\n\n확인 누르면 보고서(HTML)와 엑셀이 자동으로 열려요.'
    # 안내창을 먼저 띄우고 나서 파일을 엶. 반대로 하면 브라우저/엑셀이 화면 앞을
    # 차지해서 위의 경고 문구(스펙이탈·확인필요 등)를 못 보고 지나칠 수 있음.
    messagebox.showinfo('완료', msg)
    open_saved_files([report_path, save_path])


if __name__ == '__main__':
    # 콘솔 출력이 '깨진 글자 하나' 때문에 통째로 죽으면 안 됨. 윈도우 콘솔은
    # 보통 cp949인데, 여기 로그 문구에 쓰는 '—'(줄표) 같은 일부 기호는
    # cp949에 없어서 UnicodeEncodeError로 프로그램이 죽었음(2026-08-07 실측 —
    # process_folder는 끝까지 정상 실행됐는데 그 결과를 print(log)하다가 죽어서
    # 엑셀 저장 전에 멈춤). 진단용 콘솔 출력이 결과(엑셀/HTML)는 아니므로,
    # 못 그리는 글자는 '?'로 바뀌고 넘어가면 됨 — 문구 하나하나를 계속
    # 골라내는 것보다 이게 근본적인 해결책.
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(errors='replace')
    main()
